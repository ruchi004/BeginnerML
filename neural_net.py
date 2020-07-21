import math
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from numpy import array
from scipy.special import expit
from collections import defaultdict
from sklearn.preprocessing import LabelEncoder

dataset = pd.read_csv("C:\\Users\\RUCHI\\Desktop\\train.csv")

features = ['Pclass' , 'Age' , 'Sex']
X = dataset[features]

dataset['Survived'] = dataset['Survived']
y = dataset[['Survived']]

label_encoder = LabelEncoder()
transform_sex = array(X['Sex'])
fit_sex = label_encoder.fit_transform(transform_sex)
X['Sex'] = fit_sex


def para(dim):
	W = np.random.randn(dim,1)*0.01
	b = 0
	return W , b
W , b = para(3)

def propagate(W , b , X , Y):
	m = X.shape[1]

	Z = np.add(X.dot(W),b)
	A = expit(Z.values)
	
	Y = Y.to_numpy()
	cost = sum(A-Y)

	dw = (np.dot(X.T , np.subtract(A,y)))/m
	db = np.sum(np.subtract(A,y))/m

	cost = np.squeeze(cost)
	grads = {
			"dw" : dw ,
			"db" : db }

	return grads , cost , A , Z

def optimize(W , b , X , y , iter , rate):
	costs = []
	for i in range(iter):
		grads , cost , A , Z = propagate(W , b , X , y)

		dw = grads["dw"]
		db = grads["db"]

		W = W - dw*rate
		b = b - db*rate

		costs.append(cost)
		if i % 20 == 0:
			print("Cost after itr {} is : ".format(i) , cost , '\n\n')

	params = {
				"W" : W ,
				"b" : b	}

	grads = {"dw" : dw,
			 "db" : db
			}
	return params , grads , costs

# We will check with multiple combination of epochs and learning rate to see which combination fits better
# epochs = [100 , 200 , 300 , 400 , 500]
# l_rate = [0.001 , 0.01 , 0.02 , 0.03 , 0.05]

# hyperparameter_dict = defaultdict(list)

# min_cost = 1000
# i = 1
# for epoch in epochs:
# 	for lr in l_rate:
# 		print("For %s , %s " %(epoch , lr))
# 		params , grads , costs = optimize(W , b , X , y , epoch , lr)
# 		if min_cost>min(costs):
# 			min_cost = int(min(costs))
# 			hyperparameter_dict[str(i)].append([min_cost , epoch , lr])
# 			i+=1

# print(hyperparameter_dict)

# For this particular problem we found that epoch = 100 , and lr = 0.01 gives us the best combination. Hence we will train accordingly.

params , grads , costs = optimize(W , b , X , y , 100 , 0.01)


test_dataset = pd.read_csv("C:\\Users\\RUCHI\\Desktop\\test.csv")

X_test = test_dataset[features]

transform_sex = array(X_test['Sex'])
fit_sex = label_encoder.fit_transform(transform_sex)
X_test['Sex'] = fit_sex

p_id = test_dataset['PassengerId']

wb = Workbook()
sheet =  wb.active
sheet.title = "Prediction"

c1 = sheet.cell(row = 1, column = 1)
c1.value = 'PassengerId'
c2 = sheet.cell(row = 1, column = 2)
c2.value = 'Survived'

def predict(X_test , params):
	W = params['W']
	b = params['b']

	m = X_test.shape[1]

	Z = np.add(X_test.dot(W),b)
	A = expit(Z.values)
	an = [0]*X_test.shape[0]
	print(A)

	for index , ii in enumerate(A):
		for j in ii:
			if j!='nan' and j>0.5:
				an[index] = 1
	return an

Y_pred = predict(X_test , params)
print(Y_pred)
for ii in range(2 , len(Y_pred)+2):
	c3 = sheet.cell(row = ii , column = 1)
	c3.value = p_id[ii-2]
	c4 = sheet.cell(row = ii , column = 2)
	c4.value = Y_pred[ii-2]

wb.save("C:\\Users\\Ruchi\\Desktop\\predicted.xlsx")
