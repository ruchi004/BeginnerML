import pandas as pd
from numpy import array
import openpyxl
from openpyxl import Workbook
from sklearn.preprocessing import LabelEncoder
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, confusion_matrix , accuracy_score

dataset = pd.read_csv("C:\\Users\\RUCHI\\Desktop\\train.csv")
test_data = pd.read_csv("C:\\Users\\RUCHI\\Desktop\\test.csv")

features = ['Pclass' , 'Age' , 'Sex']
X = dataset[features]
y = dataset[['Survived']]

values = X['Age'].mean()
X['Age'] = X['Age'].fillna(values)

label_encoder = LabelEncoder()
transform_sex = array(X['Sex'])
fit_sex = label_encoder.fit_transform(transform_sex)
X['Sex'] = fit_sex

X_test = test_data[features]

valus = X_test['Age'].mean()
X_test['Age'] = X_test['Age'].fillna(valus)

label_encoder = LabelEncoder()
transform_sex = array(X_test['Sex'])
fit_sex = label_encoder.fit_transform(transform_sex)
X_test['Sex'] = fit_sex

classifier = DecisionTreeClassifier()
classifier.fit(X , y)

y_pred = classifier.predict(X_test)

p_id = test_data['PassengerId']

wb = Workbook()
sheet =  wb.active
sheet.title = "Prediction"

c1 = sheet.cell(row = 1, column = 1)
c1.value = 'PassengerId'
c2 = sheet.cell(row = 1, column = 2)
c2.value = 'Survived'

for ii in range(2 , len(y_pred)+2):
	c3 = sheet.cell(row = ii , column = 1)
	c3.value = p_id[ii-2]
	c4 = sheet.cell(row = ii , column = 2)
	c4.value = y_pred[ii-2]

wb.save("C:\\Users\\Ruchi\\Desktop\\titan.csv")

# Precision = tp/tp+fp (indicates low # of fp)
# Recall = tp/tp+fn (indicates low # of fn)
# F-measures = 2*R*P/R+P (F-measure which uses Harmonic Mean in place of Arithmetic Mean as it punishes the extreme values more.)
# Accuracy = tp + tn / tp + tn +fp +fn
